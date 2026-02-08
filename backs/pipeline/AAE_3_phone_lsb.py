import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment

# -------------------------
# Utilities
# -------------------------
def to_pascal_case(text):
    if pd.isna(text):
        return text
    return ' '.join(word.capitalize() for word in str(text).split())

def normalize_address(addr):
    if pd.isna(addr):
        return ""
    addr = str(addr).strip().lower()
    addr = re.sub(r"\s+", " ", addr)
    return addr

def ensure_folder(folder):
    if not os.path.exists(folder):
        os.makedirs(folder)

def extract_email_column(df):
    email_columns = [col for col in df.columns if 'email' in str(col).lower()]
    if email_columns:
        email_col = email_columns[0]
        return email_col, df[email_col].copy()
    return None, pd.Series([pd.NA] * len(df))

def move_email_before_list(df):
    if 'Email' in df.columns and 'List' in df.columns:
        cols = list(df.columns)
        cols.remove('Email')
        idx = cols.index('List')
        cols.insert(idx, 'Email')
        df = df[cols]
    return df

# -------------------------
# STEP 01: Clean + standardize
# -------------------------
def step_01_clean_and_standardize(df, list_name):
    _, email_data = extract_email_column(df)

    if "Owner Name" in df.columns:
        df["Owner Name"] = df["Owner Name"].astype(str).str.strip()
        df["First Name"] = df["Owner Name"].apply(lambda x: x.split(" ", 1)[0] if x else "")
        df["Last Name"] = df["Owner Name"].apply(lambda x: x.split(" ", 1)[1] if " " in x else "")

    cleaned = df.rename(columns={
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'associated_property_address_line_1': 'Property Address',
        'associated_property_address_city': 'Property City',
        'associated_property_address_state': 'Property State',
        'associated_property_address_zipcode': 'Property Zip',
        'primary_mailing_address': 'Mailing Address',
        'primary_mailing_city': 'Mailing City',
        'primary_mailing_state': 'Mailing State',
        'primary_mailing_zip': 'Mailing Zip',
        'phone_1': 'Phone1',
        'phone_2': 'Phone2',
        'phone_3': 'Phone3',
        'phone_1_type': 'Type1',
        'phone_2_type': 'Type2',
        'phone_3_type': 'Type3'
    })

    if 'Mailing Address' in cleaned.columns:
        cleaned['Mailing Address'] = cleaned['Mailing Address'].apply(to_pascal_case)
    if 'Mailing City' in cleaned.columns:
        cleaned['Mailing City'] = cleaned['Mailing City'].apply(to_pascal_case)

    cols_to_keep = [
        'First Name', 'Last Name',
        'Property Address', 'Property City', 'Property State', 'Property Zip',
        'Mailing Address', 'Mailing City', 'Mailing State', 'Mailing Zip',
        'Phone1', 'Type1', 'Phone2', 'Type2', 'Phone3', 'Type3'
    ]
    for c in cols_to_keep:
        if c not in cleaned.columns:
            cleaned[c] = pd.NA

    cleaned = cleaned[cols_to_keep]
    cleaned['Email'] = email_data.reset_index(drop=True) if not email_data.empty else pd.NA
    cleaned['List'] = list_name

    return move_email_before_list(cleaned)

# -------------------------
# STEP 02: Remove phones (2BSkip)
# -------------------------
def step_02_remove_phones(df):
    df2 = df.drop(columns=['Phone1','Type1','Phone2','Type2','Phone3','Type3','Email'], errors='ignore')
    return move_email_before_list(df2)

# -------------------------
# STEP 03: Dedupe + No Hit + CC Ready
# -------------------------
def step_03_dedupe_and_cleanup(df, list_name, output_folder=None):
    df = df.drop_duplicates()
    df['normalized_address'] = df['Property Address'].apply(normalize_address)
    df = df.drop_duplicates(subset=['normalized_address'])
    df = df.drop(columns=['normalized_address'], errors='ignore')

    # --- No Hit records ---
    mask_no_phones = df[['Phone1','Phone2','Phone3']].isna().all(axis=1)
    df_no_hit = df[mask_no_phones].copy()

    if not df_no_hit.empty and output_folder:
        nohit_folder = os.path.join(output_folder, "No Hit")
        ensure_folder(nohit_folder)
        df_no_hit = move_email_before_list(df_no_hit)
        save_to_folder(df_no_hit, nohit_folder, list_name)

    # --- Keep only rows with at least one phone ---
    df = df[~mask_no_phones].copy()

    # --- Ensure Phone4 & Phone5 exist ---
    for col in ["Phone4", "Phone5"]:
        if col not in df.columns:
            df[col] = pd.NA

    # --- Reorder: keep Phone1–Phone5 in sequence ---
    base_cols = [
        "First Name", "Last Name",
        "Property Address", "Property City", "Property State", "Property Zip",
        "Mailing Address", "Mailing City", "Mailing State", "Mailing Zip",
        "Phone1", "Phone2", "Phone3", "Phone4", "Phone5"
    ]
    other_cols = [c for c in df.columns if c not in base_cols]
    df = df[[c for c in base_cols if c in df.columns] + other_cols]

    return move_email_before_list(df)

# -------------------------
# STEP 04: Remove landlines
# -------------------------
def step_04_remove_landlines(df):
    for i in range(1,4):
        pcol, tcol = f"Phone{i}", f"Type{i}"
        if pcol not in df.columns: df[pcol] = pd.NA
        if tcol not in df.columns: df[tcol] = pd.NA
        mask = df[tcol].astype(str).str.strip().str.lower().eq("landline")
        df.loc[mask, pcol] = pd.NA

    df = df.drop(columns=['Type1','Type2','Type3'], errors='ignore')
    df = df[df[['Phone1','Phone2','Phone3']].notna().any(axis=1)].copy()
    df = df.drop(columns=['Phone4','Phone5'], errors='ignore')

    mailing_cols = ['Mailing Address','Mailing City','Mailing State','Mailing Zip']
    for c in mailing_cols:
        if c not in df.columns:
            df[c] = pd.NA

    return move_email_before_list(df)

# -------------------------
# STEP 05: Reshape for GHL
# -------------------------
def step_05_reshape(df):
    def extract_numeric(phone):
        if pd.isna(phone): 
            return None
        phone = re.sub(r'\D','',str(phone))
        if phone.endswith("0"): 
            phone = phone[:-1]
        return int(phone) if phone.isdigit() else None

    out_rows = []
    for _, row in df.iterrows():
        base = {
            "First Name": row.get("First Name", ""),
            "Last Name": row.get("Last Name", ""),
            "Property Address": row.get("Property Address", ""),
            "Property City": row.get("Property City", ""),
            "Property State": row.get("Property State", ""),
            "Property Zip": row.get("Property Zip", ""),
            "Mailing Address": row.get("Mailing Address", ""),
            "Mailing City": row.get("Mailing City", ""),
            "Mailing State": row.get("Mailing State", ""),
            "Mailing Zip": row.get("Mailing Zip", "")
        }
        phones = [extract_numeric(row.get(f"Phone{i}")) for i in range(1, 4)]
        phones = [p for p in phones if p]
        if not phones:
            out_rows.append({**base, "Phone": pd.NA, "Email": row.get("Email", ""), "List": row.get("List", "")})
        else:
            for p in phones:
                out_rows.append({**base, "Phone": p, "Email": row.get("Email", ""), "List": row.get("List", "")})

    df_out = pd.DataFrame(out_rows)

    # Drop duplicate phones
    if "Phone" in df_out.columns:
        df_out = df_out.drop_duplicates(subset=["Phone"], keep="first")

    # Reorder columns: Phone before Email, List at the very end
    cols = list(df_out.columns)
    if "Phone" in cols and "Email" in cols:
        cols.remove("Phone")
        email_idx = cols.index("Email")
        cols.insert(email_idx, "Phone")
    if "List" in cols:
        cols.remove("List")
        cols.append("List")
    df_out = df_out[cols]

    return df_out

# -------------------------
# SAVE helper (clean headers)
# -------------------------
def save_to_folder(df, folder, list_name, suffix=""):
    ensure_folder(folder)
    path = os.path.join(folder, f"{list_name}{suffix}.xlsx")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.fillna('').to_excel(writer, index=False, sheet_name="Sheet1", header=True)

    wb = load_workbook(path)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=11, bold=False, color="000000")
        cell.border = Border()
        cell.fill = PatternFill()
        cell.alignment = Alignment(horizontal="left", vertical="center")
    wb.save(path)

    print(f"✅ Saved: {path}")
    return path

# -------------------------
# RUN pipeline
# -------------------------
def run_aae_pipeline(input_path, list_name, output_folder, is_step01_file=False):
    indiv_out = os.path.join(output_folder, list_name)
    ensure_folder(indiv_out)
    step01_folder = os.path.join(indiv_out,"SkipTraced")

    if is_step01_file:
        df01 = pd.read_excel(input_path)
    else:
        if input_path.endswith('.csv'): 
            df_raw = pd.read_csv(input_path)
        else: 
            df_raw = pd.read_excel(input_path)
        df01 = step_01_clean_and_standardize(df_raw, list_name)
    path01 = save_to_folder(df01, step01_folder, list_name)

    df02 = step_02_remove_phones(df01)
    path02 = save_to_folder(df02, os.path.join(indiv_out,"2BSkip"), list_name)

    df03 = step_03_dedupe_and_cleanup(df01, list_name, indiv_out)
    path03 = save_to_folder(df03, os.path.join(indiv_out,"CC Ready"), list_name)

    df04 = step_04_remove_landlines(df03)
    path04 = save_to_folder(df04, os.path.join(indiv_out,"SC Ready"), list_name)

    df05 = step_05_reshape(df04)
    path05 = save_to_folder(df05, os.path.join(indiv_out,"GHL Ready"), list_name)

    return [path01,path02,path03,path04,path05]

# -------------------------
# Process directory
# -------------------------
def process_aae_directory(input_folder, process_step01_files=False):
    output_folder = os.path.join(input_folder,"AAE_3_Phone_LSB_Output")
    ensure_folder(output_folder)
    for file in os.listdir(input_folder):
        fpath = os.path.join(input_folder,file)
        list_name = os.path.splitext(file)[0]
        if (file.endswith(".csv") or file.endswith(".xlsx")):
            print(f"Processing {list_name}...")
            run_aae_pipeline(fpath,list_name,output_folder,process_step01_files)
    print("✅ AAE 3 Phone LSB processing complete!")

# -------------------------
# Main
# -------------------------
if __name__=="__main__":
    print("AAE 3 Phone LSB Pipeline with Email + Mailing Columns Loaded")
    print("Use: process_aae_directory('input_folder') or run_aae_pipeline(file,list,output)")
