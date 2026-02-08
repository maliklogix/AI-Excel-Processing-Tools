# utils/gen_coloumns_adder.py
import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk
import ttkbootstrap as tb
from openpyxl import load_workbook
from openpyxl.styles import Font, Border

# ---------------------------------------
# Helpers
# ---------------------------------------
def infer_column_dtype(series):
    """Infer column dtype (basic heuristic)."""
    sample = series.dropna().astype(str).head(5).tolist()
    if not sample:
        return "string"

    if all(s.isdigit() and len(s) == 5 for s in sample):
        return "zip"
    if all(s.replace("-", "").isdigit() and 7 <= len(s) <= 15 for s in sample):
        return "phone"
    try:
        pd.to_numeric(sample)
        return "numeric"
    except Exception:
        return "string"


def normalize_special_columns(df):
    """Force numeric conversion for special columns like zip and phone."""
    for col in df.columns:
        if any(x in col.lower() for x in ["Zip", "postal", "Phone"]):
            try:
                df[col] = pd.to_numeric(df[col], errors="ignore").astype("Int64")
            except Exception:
                # if not convertible, leave as string
                df[col] = df[col].astype(str)
    return df


def save_file(df, file_path, new_columns=None):
    """Save DataFrame to CSV or Excel without formatting.
       Apply dtype inference only to new columns + special normalization."""
    # Only process new columns for type inference
    if new_columns:
        for col in new_columns:
            if col in df.columns:
                dtype = infer_column_dtype(df[col])
                if dtype == "numeric":
                    df[col] = pd.to_numeric(df[col], errors="ignore")
                else:
                    df[col] = df[col].astype(str)

    # Normalize special columns (zip, phone, etc.)
    df = normalize_special_columns(df)

    # Save based on file extension
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        df.to_csv(file_path, index=False)
    else:  # Excel
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, header=True)

        # Reload and strip formatting (remove bold/borders in header)
        wb = load_workbook(file_path)
        ws = wb.active
        for cell in ws[1]:
            cell.font = Font(bold=False)
            cell.border = Border()
        wb.save(file_path)


def load_file(file_path):
    """Load CSV or Excel into DataFrame."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(file_path, dtype=str).fillna("")
    else:
        return pd.read_excel(file_path, dtype=str).fillna("")


# ---------------------------------------
# Core Logic
# ---------------------------------------
def apply_column_addition(master_file, folder, mapped_pairs, new_columns, progress_callback=None):
    """Apply new columns from master file to all Excel/CSV files in folder recursively."""
    try:
        master_df = load_file(master_file)
    except Exception as e:
        return f"❌ Failed to read master file: {e}"

    files = []
    for root_dir, _, fs in os.walk(folder):
        for fname in fs:
            if fname.lower().endswith((".xlsx", ".csv")):
                files.append(os.path.join(root_dir, fname))

    total_files = len(files)
    if total_files == 0:
        return "⚠️ No CSV/Excel files found in target folder."

    for i, file_path in enumerate(files, start=1):
        try:
            df = load_file(file_path)
        except Exception:
            continue

        added_cols = []

        # Add new columns
        for source_master_col, new_col_name, after_col, before_col in new_columns:
            if not source_master_col or not new_col_name:
                continue
            inserted = False
            for master_key_col, folder_key_col in mapped_pairs:
                if (folder_key_col in df.columns and
                    master_key_col in master_df.columns and
                    source_master_col in master_df.columns):
                    keys = master_df[master_key_col].astype(str)
                    vals = master_df[source_master_col].astype(str)
                    map_dict = dict(zip(keys, vals))
                    df[new_col_name] = df[folder_key_col].astype(str).map(map_dict).fillna("")
                    inserted = True
                    added_cols.append(new_col_name)
                    break
            if not inserted:
                df[new_col_name] = ""
                added_cols.append(new_col_name)

            # Reorder columns
            cols = list(df.columns)
            if after_col and after_col in cols:
                cols.remove(new_col_name)
                insert_at = cols.index(after_col) + 1
                cols.insert(insert_at, new_col_name)
                df = df[cols]
            if before_col and before_col in cols:
                cols.remove(new_col_name)
                insert_at = cols.index(before_col)
                cols.insert(insert_at, new_col_name)
                df = df[cols]

        # Save cleaned
        save_file(df, file_path, new_columns=added_cols)

        if progress_callback:
            progress_callback(i, total_files)

    return f"✅ Done — processed {total_files} file(s) recursively in folder."


# ---------------------------------------
# Dashboard
# ---------------------------------------
def run_column_adder(root):
    win = tb.Toplevel(root)
    win.title("Column Adder Dashboard")
    win.geometry("1200x800")
    win.configure(padx=12, pady=12)

    master_var = tk.StringVar()
    folder_var = tk.StringVar()
    status_var = tk.StringVar(value="Ready")

    master_cols = []
    target_cols = []

    # ---------- Top selectors ----------
    top_frame = tb.Frame(win)
    top_frame.pack(fill="x", pady=6)

    tb.Label(top_frame, text="Master File (CSV/Excel):").pack(side="left")
    tb.Entry(top_frame, textvariable=master_var, width=60).pack(side="left", padx=6)
    tb.Button(top_frame, text="Browse", bootstyle="secondary",
              command=lambda: select_master()).pack(side="left", padx=6)

    tb.Label(top_frame, text="Target Folder:", padding=(20,0)).pack(side="left")
    tb.Entry(top_frame, textvariable=folder_var, width=50).pack(side="left", padx=6)
    tb.Button(top_frame, text="Browse", bootstyle="secondary",
              command=lambda: select_folder()).pack(side="left", padx=6)

    # ---------- Mapping section ----------
    mapping_frame = tb.Frame(win)
    mapping_frame.pack(fill="both", expand=False, pady=8)

    lbl_frame = tb.Frame(mapping_frame)
    lbl_frame.pack(fill="x")
    tb.Label(lbl_frame, text="Master Columns").grid(row=0, column=0, padx=40)
    tb.Label(lbl_frame, text="Target Columns").grid(row=0, column=1, padx=40)
    tb.Label(lbl_frame, text="Mapped Columns").grid(row=0, column=2, padx=40)

    lists_frame = tb.Frame(mapping_frame)
    lists_frame.pack(fill="both", expand=False, pady=6)

    left_list = tk.Listbox(lists_frame, height=14, width=36, exportselection=False)
    right_list = tk.Listbox(lists_frame, height=14, width=36, exportselection=False)
    mapped_list = tk.Listbox(lists_frame, height=14, width=50, exportselection=False)

    left_list.grid(row=0, column=0, padx=12, sticky="n")
    right_list.grid(row=0, column=1, padx=12, sticky="n")
    mapped_list.grid(row=0, column=2, padx=12, sticky="n")

    btns_frame = tb.Frame(lists_frame)
    btns_frame.grid(row=0, column=3, padx=6, sticky="n")

    def add_map():
        l_idx = left_list.curselection()
        r_idx = right_list.curselection()
        if not l_idx or not r_idx:
            return
        l = left_list.get(l_idx)
        r = right_list.get(r_idx)
        mapped_list.insert("end", f"{l} == {r}")

    def remove_map():
        sel = mapped_list.curselection()
        if sel:
            mapped_list.delete(sel)

    tb.Button(btns_frame, text="➕ Map Selected", bootstyle="primary",
              width=16, command=add_map).pack(pady=6)
    tb.Button(btns_frame, text="🗑 Remove Mapping", bootstyle="danger",
              width=16, command=remove_map).pack(pady=6)

    # ---------- New Columns section ----------
    tb.Label(win, text="New Columns — fetch from Master and add to target (Insert After / Insert Before).").pack(anchor="w", pady=(10,2))
    newcols_frame = tb.Frame(win)
    newcols_frame.pack(fill="x", pady=6)

    newcol_rows = []

    def add_newcol_row():
        row = tb.Frame(newcols_frame)
        row.pack(fill="x", pady=4)

        src_dd = ttk.Combobox(row, values=master_cols, width=28)
        src_dd.pack(side="left", padx=6)
        newname_entry = ttk.Entry(row, width=24)
        newname_entry.pack(side="left", padx=6)
        after_dd = ttk.Combobox(row, values=target_cols, width=28)
        after_dd.pack(side="left", padx=6)
        before_dd = ttk.Combobox(row, values=target_cols, width=28)
        before_dd.pack(side="left", padx=6)

        remove_btn = tb.Button(row, text="Remove", bootstyle="danger",
                               command=lambda: (row.destroy(), newcol_rows.remove((src_dd, newname_entry, after_dd, before_dd))))
        remove_btn.pack(side="left", padx=6)

        newcol_rows.append((src_dd, newname_entry, after_dd, before_dd))

    tb.Button(newcols_frame, text="➕ Add New Column", bootstyle="success",
              command=add_newcol_row).pack(anchor="w", pady=4)
    add_newcol_row()

    # ---------- Progress Bar ----------
    progress = ttk.Progressbar(win, orient="horizontal", mode="determinate")
    progress.pack(fill="x", pady=10)

    # ---------- Run & Status ----------
    def run_action():
        mapped_pairs = []
        for i in range(mapped_list.size()):
            val = mapped_list.get(i)
            if "==" in val:
                a, b = val.split("==", 1)
                mapped_pairs.append((a.strip(), b.strip()))

        new_columns = []
        for src_dd, newname_ent, after_dd, before_dd in newcol_rows:
            s = src_dd.get().strip()
            n = newname_ent.get().strip()
            a = after_dd.get().strip()
            b = before_dd.get().strip()
            if s and n:
                new_columns.append((s, n, a, b))

        if not master_var.get() or not folder_var.get():
            status_var.set("⚠️ Please choose Master file and Target folder.")
            return
        if not mapped_pairs:
            status_var.set("⚠️ Please create at least one mapping (Master == Target).")
            return
        if not new_columns:
            status_var.set("⚠️ Please add at least one new column spec.")
            return

        def update_progress(done, total):
            progress["maximum"] = total
            progress["value"] = done
            win.update_idletasks()

        status_var.set("⏳ Processing...")
        win.update_idletasks()
        result = apply_column_addition(master_var.get(), folder_var.get(),
                                       mapped_pairs, new_columns, progress_callback=update_progress)
        status_var.set(result)

    tb.Button(win, text="▶ Run Column Addition", bootstyle="primary",
              width=24, command=run_action).pack(pady=10)

    status_bar = tb.Frame(win)
    status_bar.pack(fill="x", pady=(4,0))
    tb.Label(status_bar, textvariable=status_var, anchor="w").pack(fill="x")

    # ---------- Helpers ----------
    def refresh_listboxes():
        left_list.delete(0, "end")
        for c in master_cols:
            left_list.insert("end", c)
        right_list.delete(0, "end")
        for c in target_cols:
            right_list.insert("end", c)
        for src_dd, newname_ent, after_dd, before_dd in newcol_rows:
            src_dd['values'] = master_cols
            after_dd['values'] = target_cols
            before_dd['values'] = target_cols

    def select_master():
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV Files", "*.xlsx *.csv")])
        if path:
            master_var.set(path)
            try:
                cols = list(load_file(path).columns.astype(str))
            except Exception:
                cols = []
            master_cols.clear()
            master_cols.extend(cols)
            refresh_listboxes()

    def select_folder():
        path = filedialog.askdirectory()
        if path:
            folder_var.set(path)
            cols = []
            try:
                for root_dir, _, files in os.walk(path):
                    for f in files:
                        if f.lower().endswith((".xlsx", ".csv")):
                            cols = list(load_file(os.path.join(root_dir, f)).columns.astype(str))
                            break
                    if cols:
                        break
            except Exception:
                cols = []
            target_cols.clear()
            target_cols.extend(cols)
            refresh_listboxes()

    return win
