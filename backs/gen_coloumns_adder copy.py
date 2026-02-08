# utils/gen_coloumns_adder.py
import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk
import ttkbootstrap as tb
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

# ---------------------------------------
# Core Logic
# ---------------------------------------
def apply_column_addition(master_file, folder, mapped_pairs, new_columns, progress_callback=None):
    """
    master_file: path to master excel
    folder: folder containing target excel files (recursively)
    mapped_pairs: list of tuples (master_key_col, target_key_col)
    new_columns: list of tuples (source_master_col, new_column_name, insert_after_column, insert_before_column)
    progress_callback: function to update progress bar
    """
    try:
        master_df = pd.read_excel(master_file, dtype=str).fillna("")
    except Exception as e:
        return f"❌ Failed to read master file: {e}"

    # Collect all .xlsx files first for progress tracking
    all_files = []
    for root_dir, _, files in os.walk(folder):
        for fname in files:
            if fname.lower().endswith(".xlsx"):
                all_files.append(os.path.join(root_dir, fname))

    total_files = len(all_files)
    files_processed = 0

    try:
        for file_path in all_files:
            try:
                df = pd.read_excel(file_path, dtype=str).fillna("")
            except Exception:
                continue

            # Add new columns based on mappings
            for source_master_col, new_col_name, after_col, before_col in new_columns:
                if not source_master_col or not new_col_name:
                    continue
                inserted = False
                for master_key_col, folder_key_col in mapped_pairs:
                    if (folder_key_col in df.columns and
                        master_key_col in master_df.columns and
                        source_master_col in master_df.columns):
                        keys = master_df[master_key_col].astype(str)
                        vals = master_df[source_master_col].astype(str)
                        map_dict = dict(zip(keys, vals))
                        df[new_col_name] = df[folder_key_col].astype(str).map(map_dict).fillna("")
                        inserted = True
                        break
                if not inserted:
                    df[new_col_name] = ""

                # Reorder columns if requested
                cols = list(df.columns)
                if after_col and after_col in cols:
                    cols.remove(new_col_name)
                    insert_at = cols.index(after_col) + 1
                    cols.insert(insert_at, new_col_name)
                    df = df[cols]
                if before_col and before_col in cols:
                    cols.remove(new_col_name)
                    insert_at = cols.index(before_col)
                    cols.insert(insert_at, new_col_name)
                    df = df[cols]

            # Save file plain (no bold, no borders), then reformat headers
            try:
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                    df.to_excel(writer, index=False, header=True)

                # Load workbook to remove formatting + add underline
                wb = load_workbook(file_path)
                ws = wb.active

                # Remove bold & borders
                for cell in ws[1]:
                    cell.font = Font(bold=False)
                    cell.border = Border()

                # Add bottom border (horizontal line) under header row
                thin_border = Border(bottom=Side(style="thin"))
                for cell in ws[1]:
                    cell.border = thin_border

                wb.save(file_path)

            except Exception as e:
                print(f"⚠️ Could not save {file_path}: {e}")

            files_processed += 1
            if progress_callback:
                progress_callback(files_processed, total_files)

        return f"✅ Done — processed {files_processed} file(s) recursively in folder."
    except Exception as e:
        return f"❌ Error during processing: {e}"


# ---------------------------------------
# Dashboard
# ---------------------------------------
def run_column_adder(root):
    win = tb.Toplevel(root)
    win.title("Column Adder Dashboard")
    win.geometry("1200x800")
    win.configure(padx=12, pady=12)

    master_var = tk.StringVar()
    folder_var = tk.StringVar()
    status_var = tk.StringVar(value="Ready")

    master_cols = []
    target_cols = []

    # ---------- Top selectors ----------
    top_frame = tb.Frame(win)
    top_frame.pack(fill="x", pady=6)

    tb.Label(top_frame, text="Master File (Input File):").pack(side="left")
    tb.Entry(top_frame, textvariable=master_var, width=60).pack(side="left", padx=6)
    tb.Button(top_frame, text="Browse", bootstyle="secondary", command=lambda: select_master()).pack(side="left", padx=6)

    tb.Label(top_frame, text="Target Folder: Put Data Files", padding=(20,0)).pack(side="left")
    tb.Entry(top_frame, textvariable=folder_var, width=50).pack(side="left", padx=6)
    tb.Button(top_frame, text="Browse", bootstyle="secondary", command=lambda: select_folder()).pack(side="left", padx=6)

    # ---------- Mapping section ----------
    mapping_frame = tb.Frame(win)
    mapping_frame.pack(fill="both", expand=False, pady=8)

    lbl_frame = tb.Frame(mapping_frame)
    lbl_frame.pack(fill="x")
    tb.Label(lbl_frame, text="Master Columns", anchor="center").grid(row=0, column=0, padx=40)
    tb.Label(lbl_frame, text="Target Columns", anchor="center").grid(row=0, column=1, padx=40)
    tb.Label(lbl_frame, text="Mapped Columns", anchor="center").grid(row=0, column=2, padx=40)

    lists_frame = tb.Frame(mapping_frame)
    lists_frame.pack(fill="both", expand=False, pady=6)

    left_list = tk.Listbox(lists_frame, selectmode="single", height=14, width=36, exportselection=False)
    right_list = tk.Listbox(lists_frame, selectmode="single", height=14, width=36, exportselection=False)
    mapped_list = tk.Listbox(lists_frame, selectmode="single", height=14, width=50, exportselection=False)

    left_list.grid(row=0, column=0, padx=12, sticky="n")
    right_list.grid(row=0, column=1, padx=12, sticky="n")
    mapped_list.grid(row=0, column=2, padx=12, sticky="n")

    btns_frame = tb.Frame(lists_frame)
    btns_frame.grid(row=0, column=3, padx=6, sticky="n")

    def add_map():
        l_idx = left_list.curselection()
        r_idx = right_list.curselection()
        if not l_idx or not r_idx:
            return
        l = left_list.get(l_idx)
        r = right_list.get(r_idx)
        mapped_list.insert("end", f"{l} == {r}")

    def remove_map():
        sel = mapped_list.curselection()
        if sel:
            mapped_list.delete(sel)

    tb.Button(btns_frame, text="➕ Map Selected", bootstyle="primary", width=16, command=add_map).pack(pady=6)
    tb.Button(btns_frame, text="🗑 Remove Mapping", bootstyle="danger", width=16, command=remove_map).pack(pady=6)

    # ---------- New Columns section ----------
    tb.Label(win, text="New Columns — fetch from Master and add to target (Insert After / Insert Before).").pack(anchor="w", pady=(10,2))
    newcols_frame = tb.Frame(win)
    newcols_frame.pack(fill="x", pady=6)

    newcol_rows = []

    def add_newcol_row():
        row = tb.Frame(newcols_frame)
        row.pack(fill="x", pady=4)

        tb.Label(row, text="Source Column (from Master):", width=28, anchor="w").pack(side="left", padx=6)
        src_dd = ttk.Combobox(row, values=master_cols, width=28)
        src_dd.pack(side="left", padx=6)

        tb.Label(row, text="New Column Name:", width=18, anchor="w").pack(side="left", padx=6)
        newname_entry = ttk.Entry(row, width=24)
        newname_entry.pack(side="left", padx=6)

        tb.Label(row, text="Insert After Column:", width=20, anchor="w").pack(side="left", padx=6)
        after_dd = ttk.Combobox(row, values=target_cols, width=28)
        after_dd.pack(side="left", padx=6)

        tb.Label(row, text="Insert Before Column:", width=22, anchor="w").pack(side="left", padx=6)
        before_dd = ttk.Combobox(row, values=target_cols, width=28)
        before_dd.pack(side="left", padx=6)

        remove_btn = tb.Button(row, text="Remove", bootstyle="danger",
                               command=lambda: (row.destroy(), newcol_rows.remove((src_dd, newname_entry, after_dd, before_dd))))
        remove_btn.pack(side="left", padx=6)

        newcol_rows.append((src_dd, newname_entry, after_dd, before_dd))

    tb.Button(newcols_frame, text="➕ Add New Column", bootstyle="success", command=add_newcol_row).pack(anchor="w", pady=4)
    add_newcol_row()

    # ---------- Progress Bar ----------
    progress = ttk.Progressbar(win, mode="determinate", length=600)
    progress.pack(pady=8)

    # ---------- Run & Status ----------
    def run_action():
        mapped_pairs = []
        for i in range(mapped_list.size()):
            val = mapped_list.get(i)
            if "==" in val:
                a, b = val.split("==", 1)
                mapped_pairs.append((a.strip(), b.strip()))

        new_columns = []
        for src_dd, newname_ent, after_dd, before_dd in newcol_rows:
            s = src_dd.get().strip()
            n = newname_ent.get().strip()
            a = after_dd.get().strip()
            b = before_dd.get().strip()
            if s and n:
                new_columns.append((s, n, a, b))

        if not master_var.get() or not folder_var.get():
            status_var.set("⚠️ Please choose Master file and Target folder.")
            return
        if not mapped_pairs:
            status_var.set("⚠️ Please create at least one mapping (Master == Target).")
            return
        if not new_columns:
            status_var.set("⚠️ Please add at least one new column spec.")
            return

        status_var.set("⏳ Processing...")
        progress["value"] = 0
        win.update_idletasks()

        def progress_callback(done, total):
            progress["maximum"] = total
            progress["value"] = done
            win.update_idletasks()

        result = apply_column_addition(master_var.get(), folder_var.get(), mapped_pairs, new_columns, progress_callback)
        status_var.set(result)

    tb.Button(win, text="▶ Run Column Addition", bootstyle="primary", width=24, command=run_action).pack(pady=10)

    status_bar = tb.Frame(win)
    status_bar.pack(fill="x", pady=(4,0))
    tb.Label(status_bar, textvariable=status_var, anchor="w").pack(fill="x")

    # ---------- Helpers ----------
    def refresh_listboxes():
        left_list.delete(0, "end")
        for c in master_cols:
            left_list.insert("end", c)
        right_list.delete(0, "end")
        for c in target_cols:
            right_list.insert("end", c)
        for src_dd, newname_ent, after_dd, before_dd in newcol_rows:
            src_dd['values'] = master_cols
            after_dd['values'] = target_cols
            before_dd['values'] = target_cols

    def select_master():
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            master_var.set(path)
            try:
                cols = list(pd.read_excel(path, nrows=0).columns.astype(str))
            except Exception:
                cols = []
            master_cols.clear()
            master_cols.extend(cols)
            refresh_listboxes()

    def select_folder():
        path = filedialog.askdirectory()
        if path:
            folder_var.set(path)
            cols = []
            try:
                for root_dir, _, files in os.walk(path):
                    for f in files:
                        if f.lower().endswith(".xlsx"):
                            cols = list(pd.read_excel(os.path.join(root_dir, f), nrows=0).columns.astype(str))
                            break
                    if cols:
                        break
            except Exception:
                cols = []
            target_cols.clear()
            target_cols.extend(cols)
            refresh_listboxes()

    return win
